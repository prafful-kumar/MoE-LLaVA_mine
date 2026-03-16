"""
Generate Excel summary of all evaluation results.
Run: python eval_results/generate_excel.py
Reads all JSON files from eval_results/<model_variant>/<benchmark>.json
Outputs: eval_results/eval_summary.xlsx
"""

import json
import os
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

RESULTS_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(RESULTS_DIR, "eval_summary.xlsx")

# Styling
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_WHITE = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center')


def load_all_results():
    """Load all JSON result files, organized by benchmark then model."""
    results = {}  # {benchmark: {model_variant: data}}

    for variant_dir in sorted(glob.glob(os.path.join(RESULTS_DIR, "*"))):
        if not os.path.isdir(variant_dir):
            continue
        variant_name = os.path.basename(variant_dir)

        for json_file in sorted(glob.glob(os.path.join(variant_dir, "*.json"))):
            benchmark = os.path.splitext(os.path.basename(json_file))[0]
            with open(json_file) as f:
                data = json.load(f)

            if benchmark not in results:
                results[benchmark] = {}
            results[benchmark][variant_name] = data

    return results


def write_pope_sheet(wb, all_results):
    """Write POPE benchmark results with sub-categories."""
    if "pope" not in all_results:
        return

    ws = wb.create_sheet("POPE")
    pope_data = all_results["pope"]

    categories = ["random", "popular", "adversarial"]
    metrics = ["f1", "accuracy", "precision", "recall", "yes_ratio"]

    # Header row
    headers = ["Model", "Scheme", "Checkpoint", "Training Status"]
    for cat in categories:
        for metric in metrics:
            headers.append(f"{cat.capitalize()} {metric.upper()}")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Data rows
    row = 2
    for variant_name, data in sorted(pope_data.items()):
        ws.cell(row=row, column=1, value=data.get("model", variant_name)).border = THIN_BORDER
        ws.cell(row=row, column=2, value=data.get("scheme", "")).border = THIN_BORDER
        ws.cell(row=row, column=3, value=data.get("checkpoint", "")).border = THIN_BORDER
        ws.cell(row=row, column=4, value=data.get("training_status", "complete")).border = THIN_BORDER

        col = 5
        for cat in categories:
            cat_results = data.get("results", {}).get(cat, {})
            for metric in metrics:
                cell = ws.cell(row=row, column=col, value=cat_results.get(metric, ""))
                cell.alignment = CENTER
                cell.border = THIN_BORDER
                if isinstance(cell.value, float):
                    cell.number_format = '0.000'
                col += 1
        row += 1

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def flatten_results(results, prefix=""):
    """Flatten nested result dicts into key-value pairs."""
    flat = {}
    for k, v in results.items():
        key = f"{prefix}{k}" if prefix else k
        if isinstance(v, dict):
            flat.update(flatten_results(v, prefix=f"{key}/"))
        else:
            flat[key] = v
    return flat


def write_generic_sheet(wb, benchmark_name, benchmark_data):
    """Write a generic benchmark sheet (TextVQA, ScienceQA, MME, etc.)."""
    ws = wb.create_sheet(benchmark_name.upper())

    # Collect all metric keys from the first result, flattened
    sample_data = next(iter(benchmark_data.values()))
    result_keys = list(flatten_results(sample_data.get("results", {})).keys())

    headers = ["Model", "Scheme", "Checkpoint", "Training Status"] + [k.upper() for k in result_keys]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    row = 2
    for variant_name, data in sorted(benchmark_data.items()):
        ws.cell(row=row, column=1, value=data.get("model", variant_name)).border = THIN_BORDER
        ws.cell(row=row, column=2, value=data.get("scheme", "")).border = THIN_BORDER
        ws.cell(row=row, column=3, value=data.get("checkpoint", "")).border = THIN_BORDER
        ws.cell(row=row, column=4, value=data.get("training_status", "complete")).border = THIN_BORDER

        flat_results = flatten_results(data.get("results", {}))
        for col_idx, key in enumerate(result_keys, 5):
            cell = ws.cell(row=row, column=col_idx, value=flat_results.get(key, ""))
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if isinstance(cell.value, float):
                cell.number_format = '0.00'
        row += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def write_summary_sheet(wb, all_results):
    """Write a summary sheet with key metrics across all benchmarks."""
    ws = wb.active
    ws.title = "Summary"

    # Collect all model variants
    all_variants = set()
    for bench_data in all_results.values():
        all_variants.update(bench_data.keys())
    all_variants = sorted(all_variants)

    # Define which metric to show per benchmark
    key_metrics = {
        "pope": ("POPE (Adv F1)", lambda d: d.get("results", {}).get("adversarial", {}).get("f1", "")),
        "textvqa": ("TextVQA (Acc)", lambda d: d.get("results", {}).get("accuracy", "")),
        "sqa": ("ScienceQA (Acc)", lambda d: d.get("results", {}).get("accuracy", "")),
        "mme": ("MME (Total)", lambda d: d.get("results", {}).get("total", "")),
        "mmbench": ("MMBench (Acc)", lambda d: d.get("results", {}).get("accuracy", "")),
        "vqav2": ("VQAv2 (Acc)", lambda d: d.get("results", {}).get("accuracy", "")),
        "gqa": ("GQA (Acc)", lambda d: d.get("results", {}).get("accuracy", "")),
        "mmvet": ("MM-Vet (Score)", lambda d: d.get("results", {}).get("score", "")),
        "seed": ("SEED (Acc)", lambda d: d.get("results", {}).get("accuracy", "")),
    }

    # Headers
    headers = ["Model Variant"]
    available_benchmarks = []
    for bench_key, (display_name, _) in key_metrics.items():
        if bench_key in all_results:
            headers.append(display_name)
            available_benchmarks.append(bench_key)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT_WHITE
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # Data
    row = 2
    for variant in all_variants:
        ws.cell(row=row, column=1, value=variant).border = THIN_BORDER
        for col_idx, bench_key in enumerate(available_benchmarks, 2):
            bench_data = all_results[bench_key]
            if variant in bench_data:
                _, extractor = key_metrics[bench_key]
                value = extractor(bench_data[variant])
                cell = ws.cell(row=row, column=col_idx, value=value)
                if isinstance(value, float):
                    cell.number_format = '0.000'
            else:
                cell = ws.cell(row=row, column=col_idx, value="—")
            cell.alignment = CENTER
            cell.border = THIN_BORDER
        row += 1

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20


def main():
    all_results = load_all_results()

    if not all_results:
        print("No results found!")
        return

    print(f"Found benchmarks: {list(all_results.keys())}")
    for bench, variants in all_results.items():
        print(f"  {bench}: {list(variants.keys())}")

    wb = Workbook()

    # Summary sheet first
    write_summary_sheet(wb, all_results)

    # Per-benchmark sheets
    for benchmark_name, benchmark_data in sorted(all_results.items()):
        if benchmark_name == "pope":
            write_pope_sheet(wb, all_results)
        else:
            write_generic_sheet(wb, benchmark_name, benchmark_data)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
