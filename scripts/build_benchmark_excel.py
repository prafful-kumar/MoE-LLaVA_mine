"""
Build comprehensive benchmark_results.xlsx with all evaluated variants.
Sheets:
  1. Main Variants  — paper baselines + 18 main trained models
  2. Qwen Ablations — 10 hyperparameter ablations (Qwen backbone)
  3. StableLM Ablations — 10 hyperparameter ablations (StableLM backbone)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── colour palette ──────────────────────────────────────────────────────────
C_HEADER   = "1F497D"   # dark blue
C_PAPER    = "D9E1F2"   # light blue
C_PHI      = "E2EFDA"   # light green
C_QWEN     = "FFF2CC"   # light yellow
C_STABLELM = "FCE4D6"   # light orange
C_PENDING  = "D9D9D9"   # grey
C_BASELINE = "C6EFCE"   # bright green (ablation baseline)
C_BEST     = "FFEB9C"   # gold (best per backbone)

WHITE   = "FFFFFF"
BOLD    = Font(bold=True)
HEADER_FONT = Font(bold=True, color=WHITE)

def hfill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def header_row(ws, cols, row=1, fill_hex=C_HEADER):
    for c, val in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font = HEADER_FONT
        cell.fill = hfill(fill_hex)
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border()

def data_cell(ws, r, c, val, fill_hex=None, bold=False, fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    if fill_hex:
        cell.fill = hfill(fill_hex)
    if bold:
        cell.font = Font(bold=True)
    if fmt:
        cell.number_format = fmt
    cell.border = thin_border()
    cell.alignment = Alignment(horizontal="center")
    return cell

def write_data_row(ws, r, values, fill_hex=None, bold=False):
    for c, v in enumerate(values, 1):
        data_cell(ws, r, c, v, fill_hex=fill_hex, bold=bold)

def auto_width(ws, min_w=10, max_w=30):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value) if cell.value is not None else ""))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))

# ── data ─────────────────────────────────────────────────────────────────────

# POPE stored as fraction → convert to %
def p(frac):
    """fraction like 0.8897 → 88.97"""
    if frac is None or frac == "PENDING":
        return frac
    if isinstance(frac, str):
        return frac
    return round(frac * 100, 2)

def pct(s):
    """'52.11%' → 52.11"""
    if s is None or s == "PENDING":
        return s
    return float(str(s).replace("%", "").strip())

# columns for main sheet
MAIN_COLS = ["Model", "Backbone", "Init Mode", "Notes",
             "ScienceQA", "TextVQA", "GQA", "POPE (%)",
             "MME Perc.", "MME Cog.", "MME Total"]

main_data = [
    # paper baselines
    ("MoE-LLaVA-1.8B×4 (Paper)", "Qwen",     "–", "Official paper result",
     63.1,  48.0,  61.5,  87.0,  None,   None,   1291.6, C_PAPER),
    ("MoE-LLaVA-2.7B×4 (Paper)", "Phi2",     "–", "Official paper result",
     68.5,  51.4,  61.4,  86.3,  None,   None,   1423.0, C_PAPER),
    # ── Phi2 ──
    ("phi2_author",           "Phi2", "random/no_KD", "Author replica",
     71.75, 52.11, 59.44, p(0.8896907), 1378.9, 306.8, 1685.7, C_PHI),
    ("phi2_student_final",    "Phi2", "teacher_kd",   "Full student KD",
     70.76, 51.69, 60.95, p(0.8852234), 1399.2, 271.1, 1670.3, C_PHI),
    ("phi_entropy",           "Phi2", "no_teacher",   "topk_entropy_loss w=0.03",
     71.21, 51.50, 60.28, p(0.8817869), 1370.2, 308.2, 1678.4, C_PHI),
    ("phi_TS",                "Phi2", "no_teacher",   "TS Schedule",
     71.14, 52.87, 59.08, p(0.8920962), 1346.3, 315.4, 1661.7, C_PHI),
    ("phi_entropy_topk_var",  "Phi2", "no_teacher",   "topk_entropy_loss+var w=0.03",
     70.90, 51.32, 60.42, p(0.8841924), 1349.3, 286.4, 1635.8, C_PHI),
    # ── Qwen ──
    ("qwen_author",           "Qwen", "random/no_KD", "Author replica",
     63.52, 48.51, 61.96, p(0.8783505), 1325.3, 247.5, 1572.8, C_QWEN),
    ("qwen_student",          "Qwen", "teacher_kd",   "Full student KD",
     61.80, 48.08, 61.93, p(0.8790378), 1310.9, 242.9, 1553.8, C_QWEN),
    ("qwen_entropy",          "Qwen", "no_teacher",   "topk_entropy_loss w=0.03",
     61.75, 47.95, 61.88, p(0.8793814), 1323.7, 240.0, 1563.7, C_QWEN),
    ("qwen_TS",               "Qwen", "no_teacher",   "TS Schedule",
     62.72, 48.89, 61.83, p(0.8807560), 1317.2, 244.3, 1561.5, C_QWEN),
    ("qwen_entropy_w01",      "Qwen", "no_teacher",   "topk_entropy_loss w=0.1",
     62.13, 47.98, 61.91, p(0.8804124), 1329.1, 248.2, 1577.3, C_QWEN),
    ("qwen_TS_schedule",      "Qwen", "no_teacher",   "Adaptive TS Schedule",
     63.24, 48.44, 62.11, p(0.8735395), 1319.2, 247.1, 1566.4, C_QWEN),
    ("qwen_entropy_topk_var", "Qwen", "no_teacher",   "topk_entropy_loss+var w=0.03 (ablation baseline)",
     62.06, 48.00, 61.74, p(0.8814433), 1322.6, 242.9, 1565.5, C_QWEN),
    # ── StableLM ──
    ("stablelm_author",           "StableLM", "random/no_KD", "Author replica",
     59.77, 50.40, 61.97, p(0.8745704), 1322.3, 229.6, 1551.9, C_STABLELM),
    ("stablelm_student",          "StableLM", "teacher_kd",   "Full student KD",
     60.55, 49.94, 62.12, p(0.8807560), 1356.6, 241.8, 1598.4, C_STABLELM),
    ("stablelm_TS",               "StableLM", "no_teacher",   "TS Schedule",
     60.58, 50.06, 61.55, p(0.8845361), 1352.9, 224.3, 1577.2, C_STABLELM),
    ("stablelm_entropy",          "StableLM", "no_teacher",   "topk_entropy_loss w=0.03",
     59.92, 50.20, 62.02, p(0.8786942), 1347.0, 256.4, 1603.4, C_STABLELM),
    ("stablelm_entropy_topk_aux", "StableLM", "no_teacher",   "topk_entropy_loss+aux_balance",
     59.73, 49.95, 61.91, p(0.8821306), 1342.3, 242.1, 1584.4, C_STABLELM),
    ("stablelm_entropy_topk_var", "StableLM", "no_teacher",   "topk_entropy_loss+var w=0.03 (ablation baseline)",
     60.08, 50.02, 61.86, p(0.8810997), 1346.3, 249.3, 1595.6, C_STABLELM),
    ("stablelm_adaptive_entropy", "StableLM", "no_teacher",   "Margin-aware adaptive entropy gamma=2.0",
     "PENDING", "PENDING", "PENDING", "PENDING", "PENDING", "PENDING", "PENDING", C_STABLELM),
]

# ── ablation data ─────────────────────────────────────────────────────────────

ABL_COLS = ["Name", "w_ent", "lambda", "w_bal",
            "ScienceQA", "TextVQA", "GQA", "POPE (%)",
            "MME Perc.", "MME Cog.", "MME Total"]

# (name, w_ent, lam, w_bal, sqa, tvqa, gqa, pope_frac, mme_p, mme_c, mme_t, is_baseline)
qwen_abl_data = [
    # Baseline: qwen_entropy_topk_var (w_ent=0.03, lam=0.1, w_bal=0.01)
    ("qwen_entropy_topk_var [BASELINE]", 0.03, 0.1, 0.01,
     62.06, 48.00, 61.74, p(0.8814433), 1322.6, 242.9, 1565.5, True),
    # ablations
    ("qwen_abl_went_001",    0.01, 0.1,  0.01,
     62.11, 48.14, 61.63, p(0.8804124), 1329.9, 242.9, 1572.8, False),
    ("qwen_abl_went_01",     0.10, 0.1,  0.01,
     61.85, 47.98, 61.80, p(0.8780069), 1306.8, 246.8, 1553.6, False),
    ("qwen_abl_lam_001_w001",0.01, 0.01, 0.01,
     61.90, 48.25, 61.76, p(0.8814433), 1317.1, 240.7, 1557.8, False),
    ("qwen_abl_lam_001_w01", 0.10, 0.01, 0.01,
     61.83, 48.07, 61.65, p(0.8776632), 1320.1, 241.1, 1561.2, False),
    ("qwen_abl_lam_1_w001",  0.01, 1.0,  0.01,
     62.06, 48.13, 61.68, p(0.8786942), 1318.9, 236.1, 1555.0, False),
    ("qwen_abl_lam_1_w01",   0.10, 1.0,  0.01,
     62.41, 48.00, 61.75, p(0.8814433), 1327.7, 238.6, 1566.2, False),
    ("qwen_abl_wbal_0_w001", 0.01, 0.1,  0.0,
     61.80, 47.94, 61.77, p(0.8810997), 1311.6, 237.9, 1549.5, False),
    ("qwen_abl_wbal_0_w01",  0.10, 0.1,  0.0,
     62.18, 48.11, 61.87, p(0.8786942), 1319.5, 237.9, 1557.4, False),
    ("qwen_abl_wbal_01_w001",0.01, 0.1,  0.1,
     "PENDING","PENDING","PENDING","PENDING","PENDING","PENDING","PENDING", False),
    ("qwen_abl_wbal_01_w01", 0.10, 0.1,  0.1,
     61.73, 48.18, 61.76, p(0.8793814), 1333.2, 250.0, 1583.2, False),
]

stablelm_abl_data = [
    # Baseline: stablelm_entropy_topk_var (w_ent=0.03, lam=0.1, w_bal=0.01)
    ("stablelm_entropy_topk_var [BASELINE]", 0.03, 0.1, 0.01,
     60.08, 50.02, 61.86, p(0.8810997), 1346.3, 249.3, 1595.6, True),
    ("stablelm_abl_went_001",    0.01, 0.1,  0.01,
     59.84, 49.69, 62.06, p(0.8797251), 1365.5, 243.6, 1609.1, False),
    ("stablelm_abl_went_01",     0.10, 0.1,  0.01,
     60.03, 49.84, 61.99, p(0.8831615), 1363.0, 231.4, 1594.4, False),
    ("stablelm_abl_lam_001_w001",0.01, 0.01, 0.01,
     59.63, 50.07, 61.85, p(0.8841924), 1355.4, 231.1, 1586.4, False),
    ("stablelm_abl_lam_001_w01", 0.10, 0.01, 0.01,
     60.69, 49.82, 61.74, p(0.8810997), 1361.7, 246.1, 1607.8, False),
    ("stablelm_abl_lam_1_w001",  0.01, 1.0,  0.01,
     59.68, 49.82, 61.98, p(0.8855670), 1364.5, 227.1, 1591.6, False),
    ("stablelm_abl_lam_1_w01",   0.10, 1.0,  0.01,
     59.66, 50.22, 61.75, p(0.8835052), 1380.6, 263.9, 1644.5, False),
    ("stablelm_abl_wbal_0_w001", 0.01, 0.1,  0.0,
     60.48, 50.04, 62.05, p(0.8821306), 1361.4, 221.1, 1582.4, False),
    ("stablelm_abl_wbal_0_w01",  0.10, 0.1,  0.0,
     59.92, 49.82, 62.01, p(0.8793814), 1367.2, 233.6, 1600.8, False),
    ("stablelm_abl_wbal_01_w001",0.01, 0.1,  0.1,
     "PENDING","PENDING","PENDING","PENDING","PENDING","PENDING","PENDING", False),
    ("stablelm_abl_wbal_01_w01", 0.10, 0.1,  0.1,
     60.43, 50.18, 61.96, p(0.8786942), 1356.2, 234.3, 1590.5, False),
]

# ── build workbook ─────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()

# ── Sheet 1: Main Variants ─────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Main Variants"
ws1.freeze_panes = "A2"

header_row(ws1, MAIN_COLS)

backbone_color = {"Qwen": C_QWEN, "Phi2": C_PHI, "StableLM": C_STABLELM}

for r, row in enumerate(main_data, 2):
    name, backbone, init, notes, sqa, tvqa, gqa, pope, mme_p, mme_c, mme_t, fill = row
    # override fill for PENDING rows
    actual_fill = C_PENDING if sqa == "PENDING" else fill
    vals = [name, backbone, init, notes, sqa, tvqa, gqa, pope, mme_p, mme_c, mme_t]
    write_data_row(ws1, r, vals, fill_hex=actual_fill)
    # left-align the name + notes columns
    ws1.cell(r, 1).alignment = Alignment(horizontal="left")
    ws1.cell(r, 4).alignment = Alignment(horizontal="left", wrap_text=True)

# add legend below
legend_row = len(main_data) + 3
ws1.cell(legend_row, 1, "Colour legend:").font = BOLD
ws1.cell(legend_row+1, 1, "Paper baseline").fill  = hfill(C_PAPER)
ws1.cell(legend_row+2, 1, "Phi2 backbone").fill   = hfill(C_PHI)
ws1.cell(legend_row+3, 1, "Qwen backbone").fill   = hfill(C_QWEN)
ws1.cell(legend_row+4, 1, "StableLM backbone").fill = hfill(C_STABLELM)
ws1.cell(legend_row+5, 1, "Pending (eval running)").fill = hfill(C_PENDING)

auto_width(ws1)

# ── helper: write ablation sheet ──────────────────────────────────────────────

def write_ablation_sheet(ws, abl_data, title_note):
    ws.freeze_panes = "A2"

    # title note in row 1
    ws.cell(1, 1, title_note).font = Font(bold=True, size=11)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(ABL_COLS))

    header_row(ws, ABL_COLS, row=2)

    for r, row in enumerate(abl_data, 3):
        name, w_ent, lam, w_bal, sqa, tvqa, gqa, pope, mme_p, mme_c, mme_t, is_base = row
        fill = C_BASELINE if is_base else (C_PENDING if sqa == "PENDING" else WHITE)
        vals = [name, w_ent, lam, w_bal, sqa, tvqa, gqa, pope, mme_p, mme_c, mme_t]
        write_data_row(ws, r, vals, fill_hex=fill, bold=is_base)
        ws.cell(r, 1).alignment = Alignment(horizontal="left")

    # add legend
    lr = len(abl_data) + 5
    ws.cell(lr, 1, "Colour legend:").font = BOLD
    ws.cell(lr+1, 1, "Ablation baseline (already in Main Variants sheet)").fill = hfill(C_BASELINE)
    ws.cell(lr+2, 1, "Eval pending (running)").fill = hfill(C_PENDING)

    # add note about what was varied
    ws.cell(lr+4, 1, "Loss formula:  Total = w_ent × (L_leak + lambda × L_imbal) + w_bal × L_var").font = Font(italic=True)
    ws.cell(lr+5, 1, "Args:  --entropy_loss_weight (w_ent)  --imbal_lam (lambda)  --balance_loss_weight (w_bal)").font = Font(italic=True)

    auto_width(ws)

ws2 = wb.create_sheet("Qwen Ablations")
write_ablation_sheet(ws2, qwen_abl_data,
    "Qwen 1.8B Hyperparameter Ablations — topk_entropy_loss (baseline: qwen_entropy_topk_var, w_ent=0.03, lam=0.1, w_bal=0.01)")

ws3 = wb.create_sheet("StableLM Ablations")
write_ablation_sheet(ws3, stablelm_abl_data,
    "StableLM 1.6B Hyperparameter Ablations — topk_entropy_loss (baseline: stablelm_entropy_topk_var, w_ent=0.03, lam=0.1, w_bal=0.01)")

# ── save ──────────────────────────────────────────────────────────────────────
out = "benchmark_results.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"  Sheet 1 'Main Variants':       {len(main_data)} models")
print(f"  Sheet 2 'Qwen Ablations':      {len(qwen_abl_data)} variants (incl. baseline)")
print(f"  Sheet 3 'StableLM Ablations':  {len(stablelm_abl_data)} variants (incl. baseline)")
print("  PENDING rows: stablelm_adaptive_entropy, qwen_abl_wbal_01_w001, stablelm_abl_wbal_01_w001")
